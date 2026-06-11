import re
import argparse
from pathlib import Path
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIG
# ============================================================

TABLE_SETTINGS = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 3,
    "join_tolerance": 3,
    "intersection_tolerance": 3,
}

WATERMARK_LETTERS = set("FILED")

OUTPUT_FILE_NAME = "GSTR3B_Consolidated_Extract.xlsx"


# ============================================================
# CLEANING HELPERS
# ============================================================

def clean_cell(value):
    """
    Cleans extracted PDF cell text.
    Handles:
    - FILED watermark characters
    - split numbers
    - unnecessary line breaks
    """
    if value is None:
        return ""

    s = str(value).replace("\r", "\n").strip()

    if not s:
        return ""

    parts = [p.strip() for p in s.split("\n")]

    # Remove watermark letters if extracted as individual lines
    while len(parts) > 1 and parts[0] in WATERMARK_LETTERS:
        parts.pop(0)

    s = "\n".join(parts)

    # Fix split numeric values:
    # 4,76,71,026.7\n1 -> 4,76,71,026.71
    if re.fullmatch(r"[\d,\.\-\n\s]+", s):
        s = re.sub(r"\s+", "", s)
    else:
        s = re.sub(r"\s*\n\s*", " ", s)
        s = re.sub(r"[ \t]+", " ", s).strip()

    return s


def clean_table(table):
    return [[clean_cell(cell) for cell in row] for row in table]


def normalize_width(rows):
    max_cols = max((len(row) for row in rows), default=0)
    return [row + [""] * (max_cols - len(row)) for row in rows]


# ============================================================
# PDF EXTRACTION
# ============================================================

def extract_physical_tables(pdf_path):
    """
    Extracts all physical tables from a GSTR-3B PDF.
    """
    page_tables = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            tables = page.find_tables(table_settings=TABLE_SETTINGS)

            for table_index, table in enumerate(tables, start=1):
                page_tables.append({
                    "page": page_index,
                    "table_no": table_index,
                    "bbox": table.bbox,
                    "rows": clean_table(table.extract())
                })

    return page_tables


def get_rows(page_tables, page_no, table_no):
    """
    Gets rows for a specific physical table.
    """
    for table in page_tables:
        if table["page"] == page_no and table["table_no"] == table_no:
            return table["rows"]

    raise ValueError(f"Table not found: Page {page_no}, Table {table_no}")


# ============================================================
# SPECIAL TABLE HANDLING
# ============================================================

def build_6_1_table(rows):
    """
    Fixes Table 6.1 complex two-level headers.

    Parent headers:
    - Description
    - Total tax payable
    - Tax paid through ITC
    - Tax paid in cash
    - Interest paid in cash
    - Late fee paid in cash

    Child headers under Tax paid through ITC:
    - Integrated Tax
    - Central Tax
    - State/UT Tax
    - Cess
    """
    rows = normalize_width(rows)

    if len(rows) < 2:
        return rows

    rows[0] = [
        "Description",
        "Total tax payable",
        "Tax paid through ITC",
        "",
        "",
        "",
        "Tax paid in cash",
        "Interest paid in cash",
        "Late fee paid in cash"
    ]

    rows[1] = [
        "",
        "",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        "",
        "",
        ""
    ]

    return rows


def build_logical_tables(page_tables):
    """
    Maps physical PDF tables to logical GSTR-3B tables.

    This mapping is based on the structure of the GSTR-3B PDF.
    """
    return [
        {
            "sheet_name": "Period Info",
            "heading": "Form GSTR-3B — Period Information",
            "rows": get_rows(page_tables, 1, 1),
            "table_type": "normal"
        },
        {
            "sheet_name": "Registration Details",
            "heading": "Registered Person Details",
            "rows": get_rows(page_tables, 1, 2),
            "table_type": "key_value"
        },
        {
            "sheet_name": "3.1 Outward Supplies RCM",
            "heading": (
                "3.1 Details of Outward supplies and inward supplies liable "
                "to reverse charge other than those covered by Table 3.1.1"
            ),
            "rows": get_rows(page_tables, 1, 3),
            "table_type": "normal"
        },
        {
            "sheet_name": "3.1.1 Sec 9(5) Supplies",
            "heading": (
                "3.1.1 Details of Supplies notified under section 9(5) "
                "of the CGST Act, 2017 and corresponding provisions in "
                "IGST/UTGST/SGST Acts"
            ),
            "rows": get_rows(page_tables, 1, 4),
            "table_type": "normal"
        },
        {
            "sheet_name": "3.2 Inter State Supplies",
            "heading": (
                "3.2 Out of supplies made in 3.1(a) above, "
                "details of inter-state supplies made"
            ),
            "rows": get_rows(page_tables, 1, 5),
            "table_type": "normal"
        },
        {
            "sheet_name": "4 Eligible ITC",
            "heading": "4. Eligible ITC",
            "rows": get_rows(page_tables, 1, 6) + get_rows(page_tables, 2, 1),
            "table_type": "normal"
        },
        {
            "sheet_name": "5 Exempt Nil Non-GST",
            "heading": "5 Values of exempt, nil-rated and non-GST inward supplies",
            "rows": get_rows(page_tables, 2, 2),
            "table_type": "normal"
        },
        {
            "sheet_name": "5.1 Interest Late Fee",
            "heading": "5.1 Interest and Late fee for previous tax period",
            "rows": get_rows(page_tables, 2, 3),
            "table_type": "normal"
        },
        {
            "sheet_name": "6.1 Payment of Tax",
            "heading": "6.1 Payment of tax",
            "rows": build_6_1_table(get_rows(page_tables, 2, 4)),
            "table_type": "table_6_1"
        },
        {
            "sheet_name": "Breakup Tax Liability",
            "heading": "Breakup of tax liability declared for interest computation",
            "rows": get_rows(page_tables, 2, 5),
            "table_type": "normal"
        }
    ]


# ============================================================
# PROFESSIONAL STYLING
# ============================================================

class StyleBook:
    """
    Centralized professional styling.
    """

    # Corporate palette
    NAVY = "17365D"
    BLUE = "1F4E78"
    STEEL = "D9E2F3"
    LIGHT_STEEL = "EEF3F8"
    PALE_GRAY = "F7F9FB"
    MID_GRAY = "D9E1E8"
    DARK_GRAY = "404040"
    WHITE = "FFFFFF"
    GREEN = "E2F0D9"
    GOLD = "FFF2CC"

    thin_gray = Side(style="thin", color=MID_GRAY)
    medium_blue = Side(style="medium", color=BLUE)

    normal_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray
    )

    strong_border = Border(
        left=medium_blue,
        right=medium_blue,
        top=medium_blue,
        bottom=medium_blue
    )

    title_fill = PatternFill("solid", fgColor=NAVY)
    source_fill = PatternFill("solid", fgColor=LIGHT_STEEL)
    header_fill = PatternFill("solid", fgColor=STEEL)
    subheader_fill = PatternFill("solid", fgColor=LIGHT_STEEL)
    section_fill = PatternFill("solid", fgColor=PALE_GRAY)
    key_fill = PatternFill("solid", fgColor=LIGHT_STEEL)
    value_fill = PatternFill("solid", fgColor=WHITE)
    index_header_fill = PatternFill("solid", fgColor=BLUE)


def apply_cell_style(
    cell,
    fill=None,
    font=None,
    alignment=None,
    border=None
):
    if fill:
        cell.fill = fill

    if font:
        cell.font = font

    if alignment:
        cell.alignment = alignment

    if border:
        cell.border = border

def autofit_columns(ws, min_width=12, max_width=55):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for cell in column_cells:
            if cell.value is not None:
                text = str(cell.value)

                # Handle multi-line cells properly
                max_line_len = max(len(line) for line in text.split("\n"))
                max_len = max(max_len, max_line_len)

        width = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = width

# ============================================================
# WRITING SHEETS
# ============================================================

def convert_to_number(value):
    """
    Converts text numbers to float.
    Handles commas, decimals, dash values.
    """
    if value is None:
        return value

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()

    # Handle empty / dash
    if text in ["", "-"]:
        return None

    # Numeric detection
    if re.fullmatch(r"[-+]?[\\d,]*\\.?\\d+", text):
        try:
            return float(text.replace(",", ""))
        except:
            return value

    return value


def ensure_sheet(wb, sheet_name):
    """
    Creates sheet if not available.
    """
    safe_name = sheet_name[:31]

    if safe_name in wb.sheetnames:
        return wb[safe_name]

    return wb.create_sheet(safe_name)


def next_start_row(ws):
    """
    Determines next row for appending a new PDF block.
    """
    if ws.max_row == 1 and ws["A1"].value is None:
        return 1

    return ws.max_row + 3


def write_block_title(ws, start_row, max_cols, heading, source_file):
    """
    Writes professional title and source file block before each table.
    """
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=max_cols
    )

    title_cell = ws.cell(start_row, 1, heading)
    apply_cell_style(
        title_cell,
        fill=StyleBook.title_fill,
        font=Font(bold=True, color=StyleBook.WHITE, size=12),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=StyleBook.strong_border
    )
    ws.row_dimensions[start_row].height = 32

    ws.merge_cells(
        start_row=start_row + 1,
        start_column=1,
        end_row=start_row + 1,
        end_column=max_cols
    )

    source_cell = ws.cell(start_row + 1, 1, f"Source PDF: {source_file}")
    apply_cell_style(
        source_cell,
        fill=StyleBook.source_fill,
        font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=StyleBook.normal_border
    )
    ws.row_dimensions[start_row + 1].height = 22


def write_normal_table(ws, start_row, rows, heading, source_file):
    rows = normalize_width(rows)
    max_cols = max((len(row) for row in rows), default=1)

    write_block_title(ws, start_row, max_cols, heading, source_file)

    table_start = start_row + 3

    for r_idx, row in enumerate(rows, start=table_start):
        for c_idx, value in enumerate(row, start=1):

            converted_value = convert_to_number(value)
            cell = ws.cell(r_idx, c_idx, converted_value)

            # ✅ NUMBER FORMATTING
            if isinstance(converted_value, float):
                cell.number_format = '#,##0.00_);(#,##0.00)'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            apply_cell_style(
                cell,
                border=StyleBook.normal_border,
                font=Font(color=StyleBook.DARK_GRAY, size=10)
            )

            # Header row
            if r_idx == table_start:
                apply_cell_style(
                    cell,
                    fill=StyleBook.header_fill,
                    font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
                    alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                    border=StyleBook.normal_border
                )

        # Section rows
        if r_idx > table_start and row[0] and all(not x for x in row[1:]):
            for cc in range(1, max_cols + 1):
                section_cell = ws.cell(r_idx, cc)
                apply_cell_style(
                    section_cell,
                    fill=StyleBook.section_fill,
                    font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
                    alignment=Alignment(vertical="center", wrap_text=True),
                    border=StyleBook.normal_border
                )

    return table_start + len(rows)

def write_key_value_table(ws, start_row, rows, heading, source_file):
    """
    Writes key-value style tables.

    Used for Registration Details:
    GSTIN must be a row, not a header.
    """
    rows = normalize_width(rows)
    max_cols = 2

    write_block_title(ws, start_row, max_cols, heading, source_file)

    table_start = start_row + 3

    for r_idx, row in enumerate(rows, start=table_start):
        key = row[0] if len(row) > 0 else ""
        value = row[1] if len(row) > 1 else ""

        key_cell = ws.cell(r_idx, 1, key)
        value_cell = ws.cell(r_idx, 2, value)

        apply_cell_style(
            key_cell,
            fill=StyleBook.key_fill,
            font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
            alignment=Alignment(vertical="center", wrap_text=True),
            border=StyleBook.normal_border
        )

        apply_cell_style(
            value_cell,
            fill=StyleBook.value_fill,
            font=Font(color=StyleBook.DARK_GRAY, size=10),
            alignment=Alignment(vertical="center", wrap_text=True),
            border=StyleBook.normal_border
        )

    return table_start + len(rows)


def write_6_1_table(ws, start_row, rows, heading, source_file):
    rows = normalize_width(rows)
    max_cols = 9

    write_block_title(ws, start_row, max_cols, heading, source_file)

    table_start = start_row + 3

    for r_idx, row in enumerate(rows, start=table_start):
        for c_idx, value in enumerate(row, start=1):

            converted_value = convert_to_number(value)
            cell = ws.cell(r_idx, c_idx, converted_value)

            if isinstance(converted_value, float):
                cell.number_format = '#,##0.00_);(#,##0.00)'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            apply_cell_style(
                cell,
                border=StyleBook.normal_border,
                font=Font(color=StyleBook.DARK_GRAY, size=10)
            )

            # Header rows
            if r_idx in [table_start, table_start + 1]:
                apply_cell_style(
                    cell,
                    fill=StyleBook.header_fill if r_idx == table_start else StyleBook.subheader_fill,
                    font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
                    alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                    border=StyleBook.normal_border
                )

        # Section rows
        if r_idx > table_start + 1 and row[0] and all(not x for x in row[1:]):
            for cc in range(1, max_cols + 1):
                section_cell = ws.cell(r_idx, cc)
                apply_cell_style(
                    section_cell,
                    fill=StyleBook.section_fill,
                    font=Font(bold=True, color=StyleBook.DARK_GRAY, size=10),
                    alignment=Alignment(vertical="center", wrap_text=True),
                    border=StyleBook.normal_border
                )

    # ✅ KEEP YOUR MERGES (NO CHANGE)
    ws.merge_cells(start_row=table_start, start_column=1, end_row=table_start + 1, end_column=1)
    ws.merge_cells(start_row=table_start, start_column=2, end_row=table_start + 1, end_column=2)
    ws.merge_cells(start_row=table_start, start_column=3, end_row=table_start, end_column=6)
    ws.merge_cells(start_row=table_start, start_column=7, end_row=table_start + 1, end_column=7)
    ws.merge_cells(start_row=table_start, start_column=8, end_row=table_start + 1, end_column=8)
    ws.merge_cells(start_row=table_start, start_column=9, end_row=table_start + 1, end_column=9)

    return table_start + len(rows)

# ============================================================
# INDEX SHEET
# ============================================================

def create_index_sheet(wb, processed_files, failed_files):
    ws = wb.create_sheet("Index", 0)

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "GSTR-3B Consolidated Extraction Report"
    apply_cell_style(
        title,
        fill=StyleBook.title_fill,
        font=Font(bold=True, color=StyleBook.WHITE, size=14),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=StyleBook.strong_border
    )
    ws.row_dimensions[1].height = 34

    ws["A3"] = "Generated On"
    ws["B3"] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    ws["A4"] = "Total PDFs Processed"
    ws["B4"] = len(processed_files)

    ws["A5"] = "Failed PDFs"
    ws["B5"] = len(failed_files)

    for row in range(3, 6):
        apply_cell_style(
            ws.cell(row, 1),
            fill=StyleBook.key_fill,
            font=Font(bold=True, color=StyleBook.DARK_GRAY),
            border=StyleBook.normal_border
        )
        apply_cell_style(
            ws.cell(row, 2),
            fill=StyleBook.value_fill,
            font=Font(color=StyleBook.DARK_GRAY),
            border=StyleBook.normal_border
        )

    start = 8
    headers = ["Sr No", "PDF File", "Status", "Message", "Sheets Appended"]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start, col, header)
        apply_cell_style(
            cell,
            fill=StyleBook.index_header_fill,
            font=Font(bold=True, color=StyleBook.WHITE),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=StyleBook.normal_border
        )

    row_no = start + 1
    sr = 1

    for file_name in processed_files:
        values = [sr, file_name, "Success", "Extracted successfully", "All mapped sheets"]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            apply_cell_style(
                cell,
                fill=StyleBook.value_fill,
                font=Font(color=StyleBook.DARK_GRAY),
                alignment=Alignment(vertical="center", wrap_text=True),
                border=StyleBook.normal_border
            )
        row_no += 1
        sr += 1

    for file_name, error in failed_files:
        values = [sr, file_name, "Failed", str(error), "-"]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_no, col, value)
            apply_cell_style(
                cell,
                fill=PatternFill("solid", fgColor="FCE4D6"),
                font=Font(color=StyleBook.DARK_GRAY),
                alignment=Alignment(vertical="center", wrap_text=True),
                border=StyleBook.normal_border
            )
        row_no += 1
        sr += 1

    ws.freeze_panes = None
    ws.auto_filter.ref = f"A8:E{ws.max_row}"
    ws.sheet_view.showGridLines = False
    autofit_columns(ws, min_width=14, max_width=60)


# ============================================================
# MAIN CONSOLIDATION
# ============================================================

def write_logical_table_to_sheet(wb, table_spec, source_file):
    ws = ensure_sheet(wb, table_spec["sheet_name"])
    start_row = next_start_row(ws)

    table_type = table_spec["table_type"]

    if table_type == "key_value":
        write_key_value_table(
            ws=ws,
            start_row=start_row,
            rows=table_spec["rows"],
            heading=table_spec["heading"],
            source_file=source_file
        )

    elif table_type == "table_6_1":
        write_6_1_table(
            ws=ws,
            start_row=start_row,
            rows=table_spec["rows"],
            heading=table_spec["heading"],
            source_file=source_file
        )

    else:
        write_normal_table(
            ws=ws,
            start_row=start_row,
            rows=table_spec["rows"],
            heading=table_spec["heading"],
            source_file=source_file
        )

def autofit_row_heights(ws, base_height=18, line_height=15):
    for row in ws.iter_rows():
        max_lines = 1

        for cell in row:
            if cell.value:
                text = str(cell.value)

                # Count explicit line breaks
                lines = text.split("\n")
                num_lines = len(lines)

                # Estimate wrapping for long text
                max_line_len = max(len(line) for line in lines)

                col_width = ws.column_dimensions[cell.column_letter].width or 10

                # Approximate how many wrapped lines will occur
                wrapped_lines = int(max_line_len / col_width) + 1

                total_lines = max(num_lines, wrapped_lines)

                max_lines = max(max_lines, total_lines)

        # Apply row height
        ws.row_dimensions[row[0].row].height = base_height + (max_lines - 1) * line_height


def finalize_workbook_formatting(wb):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = None  # no freezing anywhere

        # Step 1: auto column width
        autofit_columns(ws, min_width=12, max_width=58)

        # Step 2: force description column wider
        if ws.max_column >= 1:
            ws.column_dimensions["A"].width = 35

        # Step 3: ensure wrap text applied everywhere
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        # Step 4: row height MUST be last
        autofit_row_heights(ws)


def consolidate_gstr3b_pdfs(input_folder, output_file):
    input_folder = Path(input_folder)

    if not input_folder.exists():
        raise FileNotFoundError(f"Folder not found: {input_folder}")

    pdf_files = sorted(input_folder.glob("*.pdf"))

    if not pdf_files:
        raise FileNotFoundError(f"No PDF files found in folder: {input_folder}")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    processed_files = []
    failed_files = []

    for pdf_file in pdf_files:
        try:
            print(f"Processing: {pdf_file.name}")

            page_tables = extract_physical_tables(pdf_file)
            logical_tables = build_logical_tables(page_tables)

            for table_spec in logical_tables:
                write_logical_table_to_sheet(
                    wb=wb,
                    table_spec=table_spec,
                    source_file=pdf_file.name
                )

            processed_files.append(pdf_file.name)

        except Exception as e:
            print(f"Failed: {pdf_file.name} -> {e}")
            failed_files.append((pdf_file.name, e))

    create_index_sheet(wb, processed_files, failed_files)
    finalize_workbook_formatting(wb)

    wb.save(output_file)

    print()
    print("Done.")
    print(f"Output file: {output_file}")
    print(f"Processed PDFs: {len(processed_files)}")
    print(f"Failed PDFs: {len(failed_files)}")


# ============================================================
# CLI ENTRY POINT
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="Batch extract GSTR-3B PDF tables into one professional Excel workbook."
    )

    parser.add_argument(
        "folder",
        help="Folder path containing GSTR-3B PDF files"
    )

    parser.add_argument(
        "-o",
        "--output",
        default=OUTPUT_FILE_NAME,
        help="Output Excel file name"
    )

    args = parser.parse_args()

    consolidate_gstr3b_pdfs(
        input_folder=args.folder,
        output_file=args.output
    )


if __name__ == "__main__":
    main()